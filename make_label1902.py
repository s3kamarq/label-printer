
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from PySimpleGUI import popup, popup_error
import print_ql

import utils
from utils import register_Teilnehmer
import pandas as pd




drucker=print_ql.etickettendrucker()

class make_label:
	def __init__(self):
		
		self.file_path = "Teilnehmerliste_Final.xlsx"
		self.logopath="mdz_Logo_Pfad.xlsx"
		###Excel Parameter
		self.UID_position = 4
		self.name_position = 1
		self.family_name_position = 2
		self.table_24_position = 6
		self.table_25_position = 7
		self.mdz = 5
		self.organisation_position = 3
		###Parameter
		self.sign_width = 696
		#self.size_qr_code = 0 #280
		#Picture Mittelstand-Digital Position
		self.qr_x_position = 400
		#self.qr_y_position = round((self.sign_width-self.size_qr_code)/2)
		self.logo_x_position = 0
		self.logo_y_position = 0
		# Picture Mittelstand-Digital Zentren
		self.mdz_x_position=0
		self.mdz_y_position=440
		#Text Positions
		self.text_left_border = 60#120  ###Changed that
		self.text_right_border =20 # New in 
		#self.title_upper_border = 0
		#self.title_lower_border = 40
		self.name_upper_border = 100#40  Changed that
		self.name_lower_border = 220#140
		self.family_name_upper_border = 170
		self.family_name_lower_border = 220
		self.organisation_upper_border = 270#270
		self.organisation_lower_border = 320 #320
		self.table_upper_border = 340
		self.table_lower_border = 370
		#self.sign_height = self.qr_x_position+self.size_qr_code
		self.sign_height_without_qr=self.qr_x_position
		self.max_fnt_size = 120	
		self.max_width_text = self.sign_width -20
		self.size_logo_height = 94
		self.size_logo_width=187
		self.size_mdz_height=94
		self.size_mdz_width= 243

		self.Druckzeit=3 # Zeit für den Drucker zu drucker vor dem es mit der GUI weiter geht




	def check_entry(self, UID):
		workbook = load_workbook(self.file_path, data_only=True)
		sheet = workbook.active
		lock = True

		for row in sheet.iter_rows():
			if str(row[self.UID_position].value).lower() == str(UID).lower():
				data_list = []
				for elements in row:
					data_list.append(str(elements.value))
				
				self.create_lable(list(data_list))
				lock = False
				register_Teilnehmer(data_list)
				break
		if lock == True:
			print(f'{UID} nicht gefunden')
			popup_error("Anmeldung nicht gefunden", 	auto_close=True, auto_close_duration = self.Druckzeit)
			
	



	def create_lable(self, data_list):

		#Read data	
		name = str(data_list[self.name_position]) + " " + str(data_list[self.family_name_position])
		#family_name = str(data_list[self.family_name_position])
		organisation = str(data_list[self.organisation_position])
		labels = [] 

		#Create sign	
		image = Image.new("RGB",(self.sign_width,self.sign_height_without_qr), color="white")
		draw = ImageDraw.Draw(image)	

		#Load Logo
		logo_image = Image.open("MD_sw.png")
		logo_image = logo_image.resize((self.size_logo_width,self.size_logo_height))
		image.paste(logo_image,(self.logo_y_position,self.logo_x_position))

		# Load Mittelstand-Digital Zentrumslogo (right side) ####################################################################
		path_list=pd.read_excel(self.logopath)
		
		##index= path_list[path_list["Zentrumsname"].str.contains(data_list[self.mdz], case=False)].index
		#index=index[0]
					
		##logo_file_path = path_list.iloc[index]['Dateiname']
		##print(logo_file_path)
		#logo_mdz= Image.open(logo_file_path) #'Logos300dpi\MD_zentrum_augsburg_CMYK_300dpi.jpg'
		logo_mdz= Image.open('Logos300dpi\MD_zentrum_augsburg_CMYK_300dpi.jpg') #'Logos300dpi\MD_zentrum_augsburg_CMYK_300dpi.jpg'
		logo_mdz= logo_mdz.resize((self.size_mdz_width, self.size_mdz_height))
		image.paste(logo_mdz,(self.mdz_y_position,self.mdz_x_position))
		
		#Draw name
		position = self.calculate_position(name,self.name_upper_border,self.name_lower_border, fnt_size = self.max_fnt_size,bolt="bolt", )
		#fnt = ImageFont.truetype("arialbd.ttf", position[0])
		fnt = ImageFont.truetype("arial.ttf", position[0])
		draw.text((position[1],position[2]),name,fill="black",font=fnt)
		
		#Draw family_name
		#position = self.calculate_position(family_name,self.family_name_upper_border,self.family_name_lower_border, fnt_size = self.max_fnt_size)
		#fnt = ImageFont.truetype("arial.ttf", position[0])
		#draw.text((position[1],position[2]),family_name,fill="black",font=fnt)
		
		#Draw organisation
		position = self.calculate_position(organisation,self.organisation_upper_border,self.organisation_lower_border,fnt_size = 55)
		fnt = ImageFont.truetype("arial.ttf", position[0])
		draw.text((position[1],position[2]),organisation,fill="black",font=fnt)
		
		image.save("namensschilder/"+data_list[self.family_name_position] + ".png")
		drucker.print_label(image)
		popup("Namensschild wird an den Drucker geschickt", 	auto_close=True, auto_close_duration = self.Druckzeit)
		#print("Finish")
		self.image = image
		#labels.append(image)

	#def get_logo(self, data_list):
	#	path_list=pd.read_excel(self.logopath)
	#	logos={}
	#	for person in data_list:
	#		if data_list[self.mdz] in path_list[0]:
	#			logo_file_path = path_list[path_list[0] == data_list[self.mdz]][1].values[0]
	#	
	#	return logo_file_path
	
	def calculate_position(self, text,upper_border,lower_border, fnt_size, bolt = "not"):

		if bolt == "bolt":
			fnt = ImageFont.truetype("arialbd.ttf",fnt_size)
		else:
			fnt = ImageFont.truetype("arial.ttf",fnt_size)
		
		max_width = self.sign_width - self.text_left_border-self.text_right_border
		max_height = lower_border - upper_border
		
		bbox = fnt.getbbox(text)
		width = bbox[2] - bbox[0]
		height = bbox[3] - bbox[1]

		while width> max_width or height > max_height:
			
			fnt_size -= 1
			if bolt == "bolt":
				fnt = ImageFont.truetype("arialbd.ttf",fnt_size)
			else:
				fnt = ImageFont.truetype("arial.ttf",fnt_size)
			bbox = fnt.getbbox(text)
		
			width = bbox[2] - bbox[0]
			height = bbox[3] - bbox[1]

		text_size = fnt_size
		x_position = self.text_left_border
		y_position = (max_height - height)/2 + upper_border


		return text_size,x_position,y_position


#für testing:
#test = make_label()
#test.check_entry('m.lundborg@wik.org')
#test.check_entry('sven.pultar@dfki.de')
	

