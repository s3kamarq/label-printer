import PySimpleGUI as sg 
from make_label1902 import make_label
import pandas as pd

sg.theme('LightGrey1')

with open('Teilnahmebedingungen.txt', 'r') as f:
    teilnahmebedingungen_text = f.read()

do_label = make_label()

import openpyxl

def add_to_excel(titel,vorname, name, organisation, email, excel_file, zentrum):
    # Öffne die Excel-Datei
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    # Finde die nächste leere Zeile
    next_row = sheet.max_row + 1
    
    # Füge die Daten in die nächste leere Zeile ein
    sheet.cell(row=next_row, column=18).value = titel
    sheet.cell(row=next_row, column=19).value = vorname
    sheet.cell(row=next_row, column=20).value = name
    sheet.cell(row=next_row, column=21).value = email
    sheet.cell(row=next_row, column=22).value = organisation
    sheet.cell(row=next_row, column=30).value = email
    sheet.cell(row=next_row, column=31).value = zentrum
  
    
    # Speichere die Änderungen
    wb.save(excel_file)

mdz= pd.read_excel('mdz_Logo_Pfad.xlsx')
mdz= mdz['Zentrumsname'].tolist()
mdz_netzwerk= ['Nein'] + mdz

def anmeldung_window():
    layout = [
        [sg.Text('Titel', size=(10, 1)),sg.Combo(['','Dr.','Prof.', 'Prof. Dr.'],key='-Titel-', size=(30,1))],
        [sg.Text('Vorname', size=(10, 1)), sg.Input(key='-Vorname-', size=(30, 1))],
        [sg.Text('Name', size=(10, 1)), sg.Input(key='-Name-', size=(30, 1))],
        [sg.Text('Organisation', size=(10, 1)), sg.Input(key='-Organisation-', size=(30, 1))],
        [sg.Text('E-Mail', size=(10, 1)), sg.Input(key='email', size=(30, 1))],
        [sg.Text('Mittelstand-Digital Netzwerk', size=(10, 1)),sg.Combo(mdz_netzwerk,key='-mdz_Netzwerk-', size=(30,1))],
        [sg.Button('Bestätigen')]
    ]

    subwindow = sg.Window('Anmeldung', layout)

    while True:
        event, values = subwindow.read()
        if event == sg.WINDOW_CLOSED:
            break
        elif event == 'Bestätigen':
            titel= values['-Titel-']
            vorname = values['-Vorname-']
            name = values['-Name-']
            organisation = values['-Organisation-']
            email = values['email']
            zentrum=values['-mdz_Netzwerk-']
            add_to_excel(titel=titel,vorname=vorname,name=name,organisation=organisation, email=email,excel_file='pretix_file.xlsx', zentrum=zentrum)
            do_label.check_entry(email)
            sg.popup('Sie haben sich erfolgreich angemeldet! Ihr Namensschild wird nun gedruckt.', auto_close=True, auto_close_duration=7)
            break

    subwindow.close()

layout = [
    [sg.Image('banner.png', size=(None, None))],
    [sg.Text(' ')],
    [sg.Text(' ')],
    [sg.Text('Willkommen!', size=(None, None), font=('Helvetica', 20), justification='center')],
    [sg.Text('Zum Einchecken bitte Ihren QR-Code an den Scanner halten.', size=(None, None), font=('Helvetica', 20), justification='center')],
    [sg.Text(' ')],
    [sg.Text(' ')],
    [sg.Text(' ')],
    [sg.Text('Ihr Ticket-Code:', size=(20, 1)), sg.Input(key='QR-Code', size=(30, 1), enable_events=True, focus=True)],  # QR-Code-Feld mit Ereignisaktivierung
    [sg.Text('Noch nicht angemeldet?', size=(20, 1), justification='left'), sg.Button('Zur Anmeldung', size=(None, None), button_color=('white', 'black'), key='Anmeldung')],  # Button "Zur Anmeldung",
]

window = sg.Window('Anmeldung zum Kongress 2024', layout, size=[850, 500], finalize=True, resizable=True)
window['QR-Code'].bind("<Return>", "_Enter")

while True:
    event, values = window.read()
    if event == sg.WINDOW_CLOSED or event == 'Exit':
        break
    elif event == 'Anmeldung':
        anmeldung_window()
    elif event == 'QR-Code' + "_Enter":  # Wenn das Eingabefeld endet mit einem Zeilenumbruch (Enter)
        email = values['QR-Code'].strip()  # Entferne den Zeilenumbruch
        do_label.check_entry(email)
        window['QR-Code'].update('')

window.close()
